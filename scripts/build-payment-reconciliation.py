import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# (rn, sheet_name, sheet_wo, sheet_recv, sheet_bal, erp_proj, erp_customer, erp_contracted, erp_recv, confidence, bucket, note)
# sheet_recv None = blank in source. erp_* None = no/ambiguous match.
rows = [
 (1,"SWAMINATHAN",434000,270000,164000,None,"3 candidates",None,None,"LOW","AMBIGUOUS",
   "Swaminathan(0267,₹0 contr) / Advocate Alwarpet(0304,3.16L) / Swaminathan-Mahalingapuram(0350,4.34L=WO exact). Pick one."),
 (2,"JAYANTHI",588000,250000,338000,"2026-27/0328","Jayanthi",588410,0,"HIGH","ERP BEHIND","Book ₹2,50,000 (ERP shows 0 received)."),
 (3,"GOPI GUDUVANCHERRY",400000,100000,300000,None,"no match",None,None,"-","NO MATCH","No project found for 'Gopi'/'Guduvancherry'. New project? Or different name?"),
 (4,"SRINIVASAN - ADYAR",565000,200000,365000,"2025-26/0032","Mr Srinivasan Adyar",565000,0,"HIGH","ERP BEHIND","WO exact. Book ₹2,00,000 (ERP shows 0)."),
 (5,"NITHYA BUILDERS",430000,300000,130000,None,"3 candidates",None,None,"LOW","AMBIGUOUS",
   "Subramaniam Nithya Builders(0116, recv 1.31L) / Nithya Akshayam(0262) / Subramaniam nithya MD house(0322,4.34L). Pick one."),
 (6,"RAMESH - BCPL",221086,128235,92851,None,"4 'Ramesh' candidates",None,None,"LOW","AMBIGUOUS",
   "None match WO 2.21L. Dr Ramesh(0080,recv1.74L) / Ramesh Babu TVS(0088) / Ramesh Babu 3.3Kw(0255) / Dr ramesh(0044). None tied to BCPL."),
 (7,"DHANDAPANI - BCPL",221086,110000,111086,None,"no match",None,None,"-","NO MATCH","No 'Dhandapani' project. New?"),
 (8,"SRIBALA - BCPL",387585,211410,176175,None,"no match",None,None,"-","NO MATCH","No 'Sribala' project. New?"),
 (9,"BCPL",495000,None,495000,None,"3 BCPL candidates",None,None,"LOW","AMBIGUOUS",
   "Sheet 'received' is BLANK. BCPL-Sridhar Rajan(0055) / BCPL-Arjun Prasath(0056) / BCPL-Anusham(0366). Which is this row?"),
 (10,"MURALI - PAMMAL",342000,None,342000,None,"no match",None,None,"-","NO MATCH","No 'Murali'/'Pammal' project. Sheet 'received' BLANK. New?"),
 (11,"LANCOR - BAGYA",238959,107827,131132,"2025-26/0037","Lancor Bagya",238958,238959,"HIGH","ERP AHEAD","ERP recv ₹2,38,959 > sheet ₹1,07,827 (fully paid in ERP). Sheet stale. No write possible."),
 (12,"LANCOR - ANANDA",238959,107827,131132,"2025-26/0163","Mr. Anandan - Nolambur?",275000,235000,"LOW","AMBIGUOUS",
   "Only 'Ananda' match is 'Anandan-Nolambur' (contr 2.75L≠WO 2.39L, recv 2.35L). Same project? Uncertain."),
 (13,"DRA - SKYLANDIS",748837,198380,550457,None,"no match",None,None,"-","NO MATCH","No 'Skylandis' project. New DRA project?"),
 (14,"DRA - TRINITY",1370979,364489,1006490,"2025-26/0012","DRA Trinity",1370979,31350,"HIGH","ERP BEHIND","WO exact. Book ₹3,33,139 (ERP recv 31,350)."),
 (15,"DRA - DOWNTOWN",2665214,435318,2229896,"2026-27/0316","DRA Downtown",2665123,0,"HIGH","ERP BEHIND","Book ₹4,35,318 (ERP shows 0)."),
 (16,"KRISHNA - EKTA",172957,41430,131527,"2025-26/0106","Krishna Ekta",172957,0,"HIGH","ERP BEHIND","WO exact. Book ₹41,430 (ERP shows 0)."),
 (17,"ZENETO",167165,123500,43665,"2025-26/0057","Zeneto Infra Pvt Ltd",167165,217165,"HIGH","ERP AHEAD","ERP recv ₹2,17,165 > sheet ₹1,23,500. Sheet stale. No write possible."),
 (18,"SATHISH",205000,100000,105000,"2025-26/0003","Mr Sathish",203285,0,"MED","ERP BEHIND","Single match (common name). Book ₹1,00,000 (ERP shows 0)."),
 (19,"4BRICKS - RBI COLONY",145000,135882,9118,"2025-26/0123","4Bricks - RBI colony",145882,145882,"HIGH","ERP AHEAD","ERP recv ₹1,45,882 > sheet ₹1,35,882 (fully paid). Sheet stale."),
 (20,"NEWRY - ADORA",800788,674493,126295,"2025-26/0038","Newry Adora",800788,824493,"HIGH","ERP AHEAD","ERP recv ₹8,24,493 > sheet ₹6,74,493. Sheet stale. No write possible."),
 (21,"GREEN VALLEY - MAHAMERU",387000,350000,37000,"2025-26/0061","Mahameru",387000,387000,"HIGH","ERP AHEAD","ERP recv ₹3,87,000 (fully paid) > sheet ₹3,50,000. Sheet stale."),
 (22,"LANCOR - LUMINA ABC",787000,340214,446786,"2025-26/0064","Lancor Lumina ABC",782406,150000,"MED","ERP BEHIND","Book ₹1,90,214 (ERP recv 1.5L). Two other 'Lumina H2' projects exist — confirm ABC is the one."),
 (23,"SHREE VIRUKSHA - SHANKARAN",632340,542000,90340,"2025-26/0028","Shankaran - Shree Viruksha 8KWp",642640,0,"HIGH","ERP BEHIND","Book ₹5,42,000 (ERP shows 0)."),
 (24,"RADIANCE - SPLENDOUR",6100000,5626867,473133,"2025-26/0118","Radiance Splendour Coimbatore",6100000,7173867,"HIGH","ERP AHEAD","ERP recv ₹71.7L > sheet ₹56.3L. Sheet stale. No write possible."),
 (25,"PRESTIGE - HILL CREST",8967223,7428225,1538998,"2025-26/0086","Prestige Hill Crest",8967231,8580584,"HIGH","ERP AHEAD","ERP recv ₹85.8L > sheet ₹74.3L. Sheet stale. No write possible."),
 (26,"S&P - COURTYARD",1200000,566729,633271,"2024-25/0041","S&P Courtyard",1150004,995303,"HIGH","ERP AHEAD","ERP recv ₹9.95L > sheet ₹5.67L. Sheet stale. No write possible."),
]

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Reconciliation"

FONT = "Calibri"
bucket_fill = {
 "ERP BEHIND": "FFF2CC",   # amber - candidate to book
 "ERP AHEAD":  "D9EAD3",   # green - already ahead, no action
 "AMBIGUOUS":  "FCE5CD",   # orange - needs a pick
 "NO MATCH":   "F4CCCC",   # red - needs identification
}
thin = Side(style="thin", color="CCCCCC")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

ws["A1"] = "Payment Follow-up — Sheet vs ERP Reconciliation (dev)"
ws["A1"].font = Font(name=FONT, bold=True, size=14)
ws["A2"] = "Source: 'Payment followup.xlsx' (26 rows) vs Shiroi ERP dev DB, 2026-06-13. Read-only — no records written. Δ to book = MAX(sheet received − ERP received, 0)."
ws["A2"].font = Font(name=FONT, italic=True, size=9, color="666666")

headers = ["#","Sheet name","Sheet WO/PO","Sheet recv","Sheet bal",
           "ERP project","ERP customer","ERP contracted","ERP recv",
           "Match","Bucket","Δ to book","Note / action needed"]
hr = 4
for c,h in enumerate(headers,1):
    cell = ws.cell(hr,c,h)
    cell.font = Font(name=FONT, bold=True, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor="434343")
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = border

r = hr+1
for (rn,nm,wo,rcv,bal,proj,cust,contr,ercv,conf,bucket,note) in rows:
    ws.cell(r,1,rn)
    ws.cell(r,2,nm)
    ws.cell(r,3,wo)
    ws.cell(r,4,rcv)
    ws.cell(r,5,bal)
    ws.cell(r,6,proj)
    ws.cell(r,7,cust)
    ws.cell(r,8,contr)
    ws.cell(r,9,ercv)
    ws.cell(r,10,conf)
    ws.cell(r,11,bucket)
    # delta to book: only meaningful when we have both numbers (computed value, not formula)
    if rcv is not None and ercv is not None:
        ws.cell(r,12,max(rcv-ercv,0))
    else:
        ws.cell(r,12,None)
    ws.cell(r,13,note)
    fill = PatternFill("solid", fgColor=bucket_fill[bucket])
    for c in range(1,14):
        cl = ws.cell(r,c)
        cl.border = border
        cl.font = Font(name=FONT, size=10)
        if c in (3,4,5,8,9,12):
            cl.number_format = "#,##0;(#,##0);-"
        if c in (10,11):
            cl.alignment = Alignment(horizontal="center")
        if c == 13:
            cl.alignment = Alignment(wrap_text=True, vertical="top")
        ws.cell(r,c).fill = fill
    r += 1

# summary block (computed values)
sum_wo = sum(rw[2] for rw in rows if rw[2] is not None)
sum_sheet_recv = sum(rw[3] for rw in rows if rw[3] is not None)
sum_erp_recv = sum(rw[8] for rw in rows if rw[8] is not None)
sum_delta = sum(max(rw[3]-rw[8],0) for rw in rows if rw[3] is not None and rw[8] is not None)
s = r+1
ws.cell(s,2,"Totals").font = Font(name=FONT, bold=True)
ws.cell(s,3,sum_wo).number_format="#,##0"
ws.cell(s,4,sum_sheet_recv).number_format="#,##0"
ws.cell(s,9,sum_erp_recv).number_format="#,##0"
ws.cell(s,12,sum_delta).number_format="#,##0"
ws.cell(s,12).font = Font(name=FONT, bold=True)
for c in range(2,14):
    ws.cell(s,c).fill = PatternFill("solid", fgColor="EFEFEF")

s2 = s+2
counts = {}
for rw in rows: counts[rw[10]] = counts.get(rw[10],0)+1
ws.cell(s2,2,"Bucket summary").font = Font(name=FONT, bold=True)
order = ["ERP BEHIND","ERP AHEAD","AMBIGUOUS","NO MATCH"]
desc = {"ERP BEHIND":"clean match, ERP < sheet — candidate to record the Δ",
        "ERP AHEAD":"ERP already ≥ sheet — sheet is stale, nothing to write (payments immutable)",
        "AMBIGUOUS":"multiple ERP candidates — need you to pick",
        "NO MATCH":"no ERP project found — need identification / may be new"}
for i,b in enumerate(order):
    rr=s2+1+i
    ws.cell(rr,2,b).fill = PatternFill("solid", fgColor=bucket_fill[b])
    ws.cell(rr,2).font = Font(name=FONT, bold=True)
    ws.cell(rr,3,counts.get(b,0))
    ws.cell(rr,4,desc[b])
    ws.cell(rr,4).font = Font(name=FONT, size=9, italic=True)

widths = [4,26,12,11,11,15,26,13,12,8,12,12,60]
for i,w in enumerate(widths,1):
    ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
ws.freeze_panes = "A5"

out = r"C:\Users\vivek\OneDrive\Desktop\Payment reconciliation - ERP vs sheet 2026-06-13.xlsx"
wb.save(out)
print("SAVED", out)
