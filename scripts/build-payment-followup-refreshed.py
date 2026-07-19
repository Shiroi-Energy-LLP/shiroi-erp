import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# rn, name, wo, received, balance, status, erp_ref, note
# received None => blank in source (not booked)
rows = [
 (1,"SWAMINATHAN",434000,270000,164000,"RECORDED","PROJ/2026-27/0350 · Swaminathan-Mahalingapuram","Recorded ₹2.70L on 13-Jun (back-fill); now matches your sheet."),
 (2,"JAYANTHI",588410,250000,338410,"RECORDED","PROJ/2026-27/0328 · Jayanthi","Recorded ₹2.50L. ERP contract ₹5,88,410 (sheet had ₹5,88,000)."),
 (3,"GOPI GUDUVANCHERRY",400000,100000,300000,"RECORDED","PROJ/2026-27/0362 · Gopi Guduvanchery","Recorded ₹1.00L. (Found despite the spelling 'Guduvanchery'.)"),
 (4,"SRINIVASAN - ADYAR",565000,200000,365000,"RECORDED","PROJ/2025-26/0032 · Mr Srinivasan Adyar","Recorded ₹2.00L."),
 (5,"NITHYA BUILDERS",434000,300000,134000,"RECORDED","PROJ/2026-27/0322 · Subramaniam nithya MD house","Recorded ₹3.00L (you mapped this to the MD house)."),
 (6,"RAMESH - BCPL",228615,128235,100380,"RECORDED","PROJ/2025-26/0255 · Mr. Ramesh Babu 3.3 Kw","Recorded ₹1.28L (mapped to Ramesh Babu 3.3Kw; ERP contract ₹2,28,615)."),
 (7,"DHANDAPANI - BCPL",221086,110000,111086,"NOT IN ERP","—","Not in ERP. Create via normal order entry, then record ₹1.10L."),
 (8,"SRIBALA - BCPL",387585,211410,176175,"NOT IN ERP","—","Not in ERP. Create via normal order entry, then record ₹2.11L."),
 (9,"BCPL",495000,None,495000,"NEEDS ID","3 candidates (Sridhar Rajan / Arjun Prasath / Anusham)","Your 'received' cell was blank. Tell me which BCPL project to reconcile."),
 (10,"MURALI - PAMMAL",342000,None,342000,"NEEDS ID","several 'Murali' projects","Your 'received' cell was blank. Identify which Murali project."),
 (11,"LANCOR - BAGYA",238958,238959,-1,"ERP AHEAD","PROJ/2025-26/0037 · Lancor Bagya","ERP fully paid (₹1 rounding). Your sheet had ₹1,07,827 — stale."),
 (12,"LANCOR - ANANDA",238959,107827,131132,"NOT IN ERP","—","Not in ERP — the 'Anandan-Nolambur' (0163) match was WRONG (per Vivek); that ₹2.35L is its own project. Lancor-Ananda is a new tower → create via order entry, then record ₹1,07,827. (Identical figures to row 11 Bagya on your sheet.)"),
 (13,"DRA - SKYLANDIS",774973,548380,226593,"ERP AHEAD","PROJ/2025-26/0110 · DRA Skylanties","Same project (spelt 'Skylanties'). ERP recv ₹5.48L > sheet ₹1.98L — stale."),
 (14,"DRA - TRINITY",1370979,364489,1006490,"RECORDED","PROJ/2025-26/0012 · DRA Trinity","Recorded ₹3.33L (ERP had ₹31,350 → now ₹3.64L)."),
 (15,"DRA - DOWNTOWN",2665123,435318,2229805,"RECORDED","PROJ/2026-27/0316 · DRA Downtown","Recorded ₹4.35L."),
 (16,"KRISHNA - EKTA",172957,41430,131527,"RECORDED","PROJ/2025-26/0106 · Krishna Ekta","Recorded ₹0.41L."),
 (17,"ZENETO",167165,217165,-50000,"ERP AHEAD","PROJ/2025-26/0057 · Zeneto Infra Pvt Ltd","ERP recv ₹2.17L (₹50K over contract). Your sheet had ₹1.24L — stale."),
 (18,"SATHISH",203285,100000,103285,"RECORDED","PROJ/2025-26/0003 · Mr Sathish","Recorded ₹1.00L."),
 (19,"4BRICKS - RBI COLONY",145882,145882,0,"ERP AHEAD","PROJ/2025-26/0123 · 4Bricks - RBI colony","ERP fully paid. Your sheet had ₹1,35,882 — stale."),
 (20,"NEWRY - ADORA",800788,824493,-23705,"ERP AHEAD","PROJ/2025-26/0038 · Newry Adora","ERP recv ₹8.24L (₹23,705 over contract). Sheet had ₹6.74L — stale."),
 (21,"GREEN VALLEY - MAHAMERU",387000,387000,0,"ERP AHEAD","PROJ/2025-26/0061 · Mahameru","ERP fully paid. Your sheet had ₹3.50L — stale."),
 (22,"LANCOR - LUMINA ABC",782406,340214,442192,"RECORDED","PROJ/2025-26/0064 · Lancor Lumina ABC","Recorded ₹1.90L (ERP had ₹1.50L → now ₹3.40L)."),
 (23,"SHREE VIRUKSHA - SHANKARAN",642640,542000,100640,"RECORDED","PROJ/2025-26/0028 · Shankaran - Shree Viruskha 8KWp","Recorded ₹5.42L."),
 (24,"RADIANCE - SPLENDOUR",6100000,7173867,-1073867,"ERP AHEAD","PROJ/2025-26/0118 · Radiance Splendour","ERP recv ₹71.7L > contract ₹61L (₹10.7L over — worth checking). Sheet had ₹56.3L."),
 (25,"PRESTIGE - HILL CREST",8967231,8580584,386647,"ERP AHEAD","PROJ/2025-26/0086 · Prestige Hill Crest","ERP recv ₹85.8L > sheet ₹74.3L — stale."),
 (26,"S&P - COURTYARD",1150004,995303,154701,"ERP AHEAD","PROJ/2024-25/0041 · S&P Courtyard","ERP recv ₹9.95L > sheet ₹5.67L — stale."),
]

status_fill = {
 "RECORDED":"FFF2CC", "ERP AHEAD":"D9EAD3", "ERP AHEAD ⚠":"D9EAD3",
 "NOT IN ERP":"F4CCCC", "NEEDS ID":"FCE5CD",
}

wb = openpyxl.Workbook(); ws = wb.active; ws.title="Payment follow-up"
FONT="Calibri"
thin=Side(style="thin",color="CCCCCC"); border=Border(left=thin,right=thin,top=thin,bottom=thin)

ws["A1"]="Payment Follow-up — reconciled with Shiroi ERP (13-Jun-2026)"
ws["A1"].font=Font(name=FONT,bold=True,size=14)
ws["A2"]=("RECEIVED / BALANCE are now live ERP figures. 12 payments (₹28,90,336) were back-filled into the ERP on 13-Jun. "
          "WO/PO = ERP contract value for matched rows (sheet value for the 4 unmatched). Dev environment.")
ws["A2"].font=Font(name=FONT,italic=True,size=9,color="666666")

headers=["#","PROJECT NAME","WO/PO VALUE","RECEIVED","BALANCE","STATUS","ERP PROJECT","NOTE"]
hr=4
for c,h in enumerate(headers,1):
    cell=ws.cell(hr,c,h)
    cell.font=Font(name=FONT,bold=True,color="FFFFFF")
    cell.fill=PatternFill("solid",fgColor="434343")
    cell.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True)
    cell.border=border

r=hr+1
for (rn,nm,wo,rcv,bal,status,ref,note) in rows:
    ws.cell(r,1,rn); ws.cell(r,2,nm); ws.cell(r,3,wo)
    ws.cell(r,4,rcv); ws.cell(r,5,bal); ws.cell(r,6,status)
    ws.cell(r,7,ref); ws.cell(r,8,note)
    fill=PatternFill("solid",fgColor=status_fill[status])
    for c in range(1,9):
        cl=ws.cell(r,c); cl.border=border; cl.fill=fill; cl.font=Font(name=FONT,size=10)
        if c in (3,4,5): cl.number_format="#,##0;(#,##0);-"
        if c==6: cl.alignment=Alignment(horizontal="center")
        if c in (7,8): cl.alignment=Alignment(wrap_text=True,vertical="top")
    r+=1

# totals
tot_wo=sum(x[2] for x in rows)
tot_rcv=sum(x[3] for x in rows if x[3] is not None)
tot_bal=sum(x[4] for x in rows)
s=r+1
ws.cell(s,2,"Totals").font=Font(name=FONT,bold=True)
ws.cell(s,3,tot_wo).number_format="#,##0"
ws.cell(s,4,tot_rcv).number_format="#,##0"
ws.cell(s,5,tot_bal).number_format="#,##0;(#,##0);-"
for c in range(2,9): ws.cell(s,c).fill=PatternFill("solid",fgColor="EFEFEF")

s2=s+2
summary=[
 ("RECORDED",12,"₹28,90,336 back-filled into ERP on 13-Jun (now match your sheet)"),
 ("ERP AHEAD",9,"ERP already ≥ your sheet — figures refreshed from ERP (sheet was stale)"),
 ("NOT IN ERP",3,"Dhandapani-BCPL, Sribala-BCPL, Lancor-Ananda — create via normal order entry, then record"),
 ("NEEDS ID",2,"BCPL, Murali-Pammal — your 'received' was blank; identify the project"),
]
ws.cell(s2,2,"Summary").font=Font(name=FONT,bold=True)
for i,(st,n,d) in enumerate(summary):
    rr=s2+1+i
    ws.cell(rr,2,st).fill=PatternFill("solid",fgColor=status_fill[st]); ws.cell(rr,2).font=Font(name=FONT,bold=True)
    ws.cell(rr,3,n); ws.cell(rr,4,d).font=Font(name=FONT,size=9,italic=True)

for i,w in enumerate([4,26,13,12,12,12,40,60],1):
    ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width=w
ws.freeze_panes="A5"

out=r"C:\Users\vivek\OneDrive\Desktop\Payment followup - reconciled 2026-06-13.xlsx"
wb.save(out); print("SAVED",out,"| recorded total check:",tot_rcv)
