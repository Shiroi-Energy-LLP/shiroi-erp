"""Read-only inspection of Manivel's project dashboard workbook.

Dumps sheet names, dimensions, full headers, and a few sample rows so we can
map the source-of-truth columns against the ERP schema. No writes anywhere.
"""
import openpyxl

PATH = r"C:\Users\vivek\Projects\shiroi-erp\scripts\data\manivel-project-dashboard-2026-06-15.xlsx"

wb = openpyxl.load_workbook(PATH, read_only=True, data_only=True)
print("SHEETS:", wb.sheetnames)

for ws in wb.worksheets:
    print("=" * 90)
    print(f"SHEET: {ws.title!r}  max_row={ws.max_row}  max_col={ws.max_column}")
    sample = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        sample.append(row)
        if i >= 5:
            break
    if not sample:
        print("  (empty)")
        continue
    header = sample[0]
    print("  HEADER:")
    for ci, h in enumerate(header):
        if h is not None and str(h).strip() != "":
            print(f"    col{ci}: {h!r}")
    print("  SAMPLE ROWS:")
    for ri, row in enumerate(sample[1:], start=1):
        cells = [("" if v is None else str(v))[:30] for v in row]
        print(f"    r{ri}: " + " | ".join(cells))
